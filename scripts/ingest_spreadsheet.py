#!/usr/bin/env python3
"""Ingest a MIP spreadsheet into the standalone atlas HTML.

The spreadsheet does not include latitude/longitude. This script preserves geo data
from the current HTML by matching schools, with city-level fallback coordinates.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

FIELDS = [
    "name", "district", "address", "city", "state", "zip", "country", "grades",
    "type", "phone", "year", "mandarin", "model", "otherLang", "website",
    "language", "lat", "lng", "approx",
]


def clean(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).replace("\xa0", " ").strip()


def cell(row: tuple[Any, ...], idx: int) -> Any:
    return row[idx] if idx < len(row) else None


def title_model(v: str) -> str:
    v = clean(v)
    return {"strand": "Strand", "whole": "Whole School", "whole school": "Whole School"}.get(v.lower(), v)


COUNTRY_ALIASES = {
    "u.k.": "United Kingdom",
    "uk": "United Kingdom",
    "usa": "USA",
    "u.s.a.": "USA",
    "united states": "USA",
}


def canonical_country(v: str) -> str:
    v = clean(v)
    return COUNTRY_ALIASES.get(v.lower(), v)


def norm(*parts: str) -> str:
    fixed = [canonical_country(p) if i == len(parts) - 1 else clean(p) for i, p in enumerate(parts)]
    return "|".join(re.sub(r"[^a-z0-9]+", " ", p.lower()).strip() for p in fixed)


def geo_key(s: dict[str, Any]) -> str:
    return norm(s.get("name", ""), s.get("address", ""), s.get("city", ""), s.get("state", ""), s.get("zip", ""), s.get("country", ""))


def geocode_query(s: dict[str, Any]) -> str:
    parts = [s.get("name", ""), s.get("address", ""), s.get("city", ""), s.get("state", ""), s.get("zip", ""), s.get("country", "")]
    return ", ".join(p for p in map(clean, parts) if p)


def load_geocode_cache(path: Path) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def write_geocode_cache(path: Path, cache: dict[str, dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(cache, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def extract_const(html: str, name: str, next_name: str) -> Any:
    m = re.search(rf"const {name} = (.*?);\nconst {next_name}\b", html, re.S)
    if not m:
        raise SystemExit(f"Could not find const {name} before const {next_name}")
    return json.loads(m.group(1))


def read_existing(html_path: Path, schools_json_path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    html = html_path.read_text(encoding="utf-8")
    try:
        schools = extract_const(html, "SCHOOLS", "WORLD")
    except SystemExit:
        schools = json.loads(schools_json_path.read_text(encoding="utf-8")) if schools_json_path.exists() else []
    labels = extract_const(html, "LABELS", "C_MANDARIN")
    return schools, labels


def base_school(language: str = "Mandarin") -> dict[str, Any]:
    return {
        "name": "", "district": "", "address": "", "city": "", "state": "", "zip": "",
        "country": "USA", "grades": "", "type": "", "phone": "", "year": "",
        "mandarin": "", "model": "", "otherLang": "", "website": "",
        "language": language, "lat": None, "lng": None, "approx": True,
    }


def parse_us(ws) -> list[dict[str, Any]]:
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not clean(row[0] if row else None):
            continue
        s = base_school("Mandarin")
        s.update({
            "name": clean(row[0]), "district": clean(row[1]), "address": clean(row[2]),
            "city": clean(row[3]), "state": clean(row[4]), "zip": clean(row[5]),
            "grades": clean(row[6]), "type": clean(row[7]), "phone": clean(row[8]),
            "year": clean(row[9]), "mandarin": clean(row[10]), "model": title_model(row[11]),
            "otherLang": clean(row[13]), "website": clean(row[14]),
        })
        out.append(s)
    return out


def parse_cantonese(ws) -> list[dict[str, Any]]:
    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not clean(row[0] if row else None):
            continue
        s = base_school("Cantonese")
        district = clean(row[2]) or clean(row[1])
        address = clean(row[3])
        s.update({
            "name": clean(row[0]), "district": district, "address": address,
            "city": clean(row[4]), "state": clean(row[5]), "zip": clean(row[6]),
            "grades": clean(row[13]) or clean(row[9]), "type": clean(row[7]), "phone": clean(row[8]),
            "year": clean(row[10]), "mandarin": "", "model": title_model(row[11]),
            "otherLang": "", "website": clean(row[14]),
        })
        out.append(s)
    return out


def parse_international(ws) -> list[dict[str, Any]]:
    out = []
    current_country = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = clean(row[0] if row else None)
        nonempty = [clean(v) for v in row if clean(v)]
        if not nonempty:
            continue
        # Country separator rows have only the country name in column A.
        if name and len(nonempty) == 1:
            current_country = name
            continue
        if not name:
            continue
        s = base_school("Mandarin")
        country = canonical_country(clean(row[6]) or current_country)
        s.update({
            "name": name, "district": clean(row[2]), "address": clean(row[3]),
            "city": clean(row[4]), "state": clean(row[5]), "zip": clean(row[7]),
            "country": country or "", "grades": clean(row[10]), "type": clean(row[8]),
            "phone": clean(cell(row, 9)), "year": clean(cell(row, 11)), "mandarin": clean(cell(row, 14)),
            "model": title_model(cell(row, 15)), "otherLang": "", "website": clean(cell(row, 12)),
        })
        out.append(s)
    return out


def geo_indexes(existing: list[dict[str, Any]]):
    exact, by_name, by_name_only, by_city = {}, {}, {}, defaultdict(list)
    for s in existing:
        if s.get("lat") is None or s.get("lng") is None:
            continue
        exact[norm(s.get("name", ""), s.get("city", ""), s.get("state", ""), s.get("country", ""))] = s
        by_name.setdefault(norm(s.get("name", ""), s.get("country", "")), s)
        by_name_only.setdefault(norm(s.get("name", "")), s)
        by_city[norm(s.get("city", ""), s.get("state", ""), s.get("country", ""))].append(s)
    return exact, by_name, by_name_only, by_city


def seed_cache_from_existing(cache: dict[str, dict[str, Any]], existing: list[dict[str, Any]]) -> None:
    for s in existing:
        if s.get("lat") is None or s.get("lng") is None:
            continue
        key = geo_key(s)
        cache.setdefault(key, {
            "lat": float(s["lat"]),
            "lng": float(s["lng"]),
            "approx": bool(s.get("approx", False)),
            "source": "existing-html",
            "query": geocode_query(s),
        })


def apply_geo_hit(s: dict[str, Any], hit: dict[str, Any]) -> None:
    s["lat"], s["lng"] = float(hit["lat"]), float(hit["lng"])
    s["approx"] = bool(hit.get("approx", False))


def attach_known_geo(schools: list[dict[str, Any]], existing: list[dict[str, Any]], cache: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    exact, by_name, by_name_only, _by_city = geo_indexes(existing)
    missing = []
    for s in schools:
        hit = cache.get(geo_key(s))
        if hit:
            apply_geo_hit(s, hit)
            continue
        hit = exact.get(norm(s["name"], s["city"], s["state"], s["country"]))
        if not hit:
            hit = by_name.get(norm(s["name"], s["country"]))
        if not hit:
            hit = by_name_only.get(norm(s["name"]))
        if hit:
            s["lat"], s["lng"], s["approx"] = hit["lat"], hit["lng"], bool(hit.get("approx", False))
            cache.setdefault(geo_key(s), {
                "lat": float(s["lat"]), "lng": float(s["lng"]), "approx": s["approx"],
                "source": "existing-html", "query": geocode_query(s),
            })
        else:
            missing.append(s)
    return missing


def apply_city_fallback(schools: list[dict[str, Any]], existing: list[dict[str, Any]]) -> list[dict[str, Any]]:
    _exact, _by_name, _by_name_only, by_city = geo_indexes(existing + [s for s in schools if s.get("lat") is not None])
    missing = []
    for s in schools:
        if s.get("lat") is not None and s.get("lng") is not None:
            continue
        city_hits = by_city.get(norm(s["city"], s["state"], s["country"]), [])
        if city_hits:
            s["lat"] = round(sum(float(x["lat"]) for x in city_hits) / len(city_hits), 4)
            s["lng"] = round(sum(float(x["lng"]) for x in city_hits) / len(city_hits), 4)
            s["approx"] = True
        else:
            missing.append(s)
    return missing


def geocode_one_nominatim(s: dict[str, Any], user_agent: str) -> dict[str, Any] | None:
    query = geocode_query(s)
    url = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode({
        "q": query,
        "format": "jsonv2",
        "limit": "1",
        "addressdetails": "0",
    })
    req = urllib.request.Request(url, headers={"User-Agent": user_agent})
    with urllib.request.urlopen(req, timeout=30) as resp:
        rows = json.loads(resp.read().decode("utf-8"))
    if not rows:
        return None
    row = rows[0]
    return {
        "lat": round(float(row["lat"]), 6),
        "lng": round(float(row["lon"]), 6),
        "approx": False,
        "source": "nominatim",
        "query": query,
        "display_name": row.get("display_name", ""),
    }


def geocode_missing(schools: list[dict[str, Any]], cache: dict[str, dict[str, Any]], user_agent: str, delay: float) -> list[dict[str, Any]]:
    still_missing = []
    candidates = [s for s in schools if s.get("lat") is None or s.get("lng") is None]
    for i, s in enumerate(candidates, 1):
        print(f"Geocoding {i}/{len(candidates)}: {s['name']} — {geocode_query(s)}", file=sys.stderr)
        try:
            hit = geocode_one_nominatim(s, user_agent)
        except Exception as e:
            print(f"  geocode error: {e}", file=sys.stderr)
            hit = None
        if hit:
            cache[geo_key(s)] = hit
            apply_geo_hit(s, hit)
            print(f"  -> {hit['lat']}, {hit['lng']} ({hit.get('display_name', '')})", file=sys.stderr)
        else:
            still_missing.append(s)
            print("  -> no result", file=sys.stderr)
        if i < len(candidates):
            time.sleep(delay)
    return still_missing


def make_city_labels(schools: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets = defaultdict(list)
    for s in schools:
        if s.get("lat") is None or s.get("lng") is None or not s.get("city"):
            continue
        buckets[(s["city"], s.get("state", ""), s.get("country", ""))].append(s)
    labels = []
    for (city, state, country), rows in buckets.items():
        labels.append({
            "name": city,
            "lat": round(sum(float(r["lat"]) for r in rows) / len(rows), 3),
            "lng": round(sum(float(r["lng"]) for r in rows) / len(rows), 3),
            "n": len(rows),
        })
    labels.sort(key=lambda x: (-x["n"], x["name"]))
    return labels[:100]


def replace_const(html: str, name: str, next_name: str, value: Any) -> str:
    payload = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    pattern = rf"const {name} = .*?;\nconst {next_name}\b"
    replacement = f"const {name} = {payload};\nconst {next_name}"
    return re.sub(pattern, lambda _m: replacement, html, count=1, flags=re.S)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("spreadsheet", type=Path)
    ap.add_argument("--html", type=Path, default=Path("Chinese-Immersion-Atlas.html"))
    ap.add_argument("--schools-json", type=Path, default=Path("data/schools.json"))
    ap.add_argument("--geocode-cache", type=Path, default=Path("data/geocodes.json"))
    ap.add_argument("--geocode-missing", dest="geocode_missing", action="store_true", default=True, help="Fetch missing coordinates from Nominatim/OpenStreetMap (default)")
    ap.add_argument("--no-geocode-missing", dest="geocode_missing", action="store_false", help="Do not fetch missing coordinates; only use cache/existing HTML")
    ap.add_argument("--geocode-delay", type=float, default=1.1, help="Seconds between Nominatim requests")
    ap.add_argument("--user-agent", default="chinese-immersion-atlas/1.0", help="User-Agent for Nominatim requests")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--allow-missing-geo", action="store_true")
    args = ap.parse_args()

    existing, labels = read_existing(args.html, args.schools_json)
    cache = load_geocode_cache(args.geocode_cache)
    seed_cache_from_existing(cache, existing)

    wb = load_workbook(args.spreadsheet, read_only=True, data_only=True)
    schools = parse_us(wb["MIP List"]) + parse_international(wb["International"]) + parse_cantonese(wb["Cantonese"])
    missing = attach_known_geo(schools, existing, cache)
    if missing and args.geocode_missing:
        geocode_missing(schools, cache, args.user_agent, args.geocode_delay)
    missing = apply_city_fallback(schools, existing)

    if missing and not args.allow_missing_geo:
        print(f"Missing geo for {len(missing)} schools; HTML not changed:", file=sys.stderr)
        for s in missing[:25]:
            print(f"- {s['name']} ({s.get('city')}, {s.get('state')}, {s.get('country')})", file=sys.stderr)
        if len(missing) > 25:
            print(f"... and {len(missing)-25} more", file=sys.stderr)
        print("Rerun with --geocode-missing, add entries to data/geocodes.json, or rerun with --allow-missing-geo.", file=sys.stderr)
        return 2

    schools = [{k: s[k] for k in FIELDS} for s in schools if s.get("lat") is not None and s.get("lng") is not None]
    labels["cities"] = make_city_labels(schools)

    print(f"Parsed {len(schools)} schools ({Counter(s['language'] for s in schools)})")
    print(f"Missing geo: {len(missing)}")
    if args.dry_run:
        return 0

    write_geocode_cache(args.geocode_cache, cache)
    args.schools_json.parent.mkdir(parents=True, exist_ok=True)
    args.schools_json.write_text(json.dumps(schools, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    html = args.html.read_text(encoding="utf-8")
    html = replace_const(html, "LABELS", "C_MANDARIN", labels)
    args.html.write_text(html, encoding="utf-8")
    print(f"Updated {args.html}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
