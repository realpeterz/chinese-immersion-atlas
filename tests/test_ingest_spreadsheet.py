import importlib.util
import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SPEC = importlib.util.spec_from_file_location("ingest_spreadsheet", ROOT / "scripts" / "ingest_spreadsheet.py")
ingest = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(ingest)


class IngestSpreadsheetTests(unittest.TestCase):
    def test_parse_us_sheet_normalizes_values(self):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "School", "School District", "Address", "City", "State", "ZIP code", "Grades",
            "Public/Charter/Private", "Phone number", "Year started", "% Mandarin time",
            "Strand or Whole School", "Charter type", "Other Immersion Languages", "Website",
        ])
        ws.append([
            " Test School ", " Test District ", " 123 Main St ", "Seattle\xa0", "WA\xa0", 98101,
            "K - 5", "Public", "555-1212", "2025-2026", "50/50", "strand", "simplified",
            "Spanish", "example.org",
        ])

        schools = ingest.parse_us(ws)

        self.assertEqual(len(schools), 1)
        self.assertEqual(schools[0]["name"], "Test School")
        self.assertEqual(schools[0]["city"], "Seattle")
        self.assertEqual(schools[0]["state"], "WA")
        self.assertEqual(schools[0]["zip"], "98101")
        self.assertEqual(schools[0]["model"], "Strand")
        self.assertEqual(schools[0]["language"], "Mandarin")

    def test_parse_international_country_headers_and_aliases(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["International schools"])
        ws.append([])
        ws.append(["U.K."])
        ws.append(["Kensington Wade School", None, None, "Fulham Palace Road", "London", None, None, None, "Private", None, "K - 8", None, "https://example.org"])

        schools = ingest.parse_international(ws)

        self.assertEqual(len(schools), 1)
        self.assertEqual(schools[0]["country"], "United Kingdom")
        self.assertEqual(schools[0]["city"], "London")
        self.assertEqual(schools[0]["type"], "Private")

    def test_attach_known_geo_prefers_cache(self):
        school = ingest.base_school()
        school.update({"name": "Test School", "address": "123 Main", "city": "Seattle", "state": "WA", "zip": "98101", "country": "USA"})
        cache = {
            ingest.geo_key(school): {
                "lat": 47.61,
                "lng": -122.33,
                "approx": False,
                "source": "test",
                "query": "Test School, 123 Main, Seattle, WA, 98101, USA",
            }
        }

        missing = ingest.attach_known_geo([school], existing=[], cache=cache)

        self.assertEqual(missing, [])
        self.assertEqual(school["lat"], 47.61)
        self.assertEqual(school["lng"], -122.33)
        self.assertFalse(school["approx"])

    def test_apply_city_fallback_marks_approximate(self):
        school = ingest.base_school()
        school.update({"name": "New School", "city": "Seattle", "state": "WA", "country": "USA"})
        existing = [{"name": "Known", "city": "Seattle", "state": "WA", "country": "USA", "lat": 47.6, "lng": -122.3, "approx": False}]

        missing = ingest.apply_city_fallback([school], existing)

        self.assertEqual(missing, [])
        self.assertEqual(school["lat"], 47.6)
        self.assertEqual(school["lng"], -122.3)
        self.assertTrue(school["approx"])

    def test_replace_const_handles_backslashes_in_json_payload(self):
        html = 'const SCHOOLS = [];' + "\n" + 'const WORLD = {"old":true};'
        schools = [{"name": "Line School", "address": "One\\nTwo"}]

        updated = ingest.replace_const(html, "SCHOOLS", "WORLD", schools)
        payload = updated.split("const SCHOOLS = ", 1)[1].split(";\nconst WORLD", 1)[0]

        self.assertEqual(json.loads(payload), schools)
        self.assertIn('const WORLD = {"old":true};', updated)

    def test_main_writes_schools_and_labels_json(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            html = tmp / "atlas.html"
            html.write_text('const WORLD = {};\nconst C_MANDARIN = "#c";\n', encoding="utf-8")
            schools_json = tmp / "schools.json"
            labels_json = tmp / "labels.json"
            geocodes_json = tmp / "geocodes.json"
            xlsx = tmp / "schools.xlsx"

            schools_json.write_text(json.dumps([{
                "name": "Test School", "district": "", "address": "123 Main St", "city": "Seattle",
                "state": "WA", "zip": "98101", "country": "USA", "grades": "", "type": "",
                "phone": "", "year": "", "mandarin": "", "model": "", "otherLang": "",
                "website": "", "language": "Mandarin", "lat": 47.61, "lng": -122.33,
                "approx": False,
            }]), encoding="utf-8")
            labels_json.write_text(json.dumps({"countries": [], "states": [], "cities": []}), encoding="utf-8")

            wb = Workbook()
            ws = wb.active
            ws.title = "MIP List"
            ws.append([
                "School", "School District", "Address", "City", "State", "ZIP code", "Grades",
                "Public/Charter/Private", "Phone number", "Year started", "% Mandarin time",
                "Strand or Whole School", "Charter type", "Other Immersion Languages", "Website",
            ])
            ws.append(["Test School", "", "123 Main St", "Seattle", "WA", "98101", "K - 5", "Public", "", "", "50/50", "strand", "", "", ""])
            wb.create_sheet("International").append(["International schools"])
            wb.create_sheet("Cantonese").append(["Cantonese immersion"])
            wb.save(xlsx)

            old_argv = sys.argv
            try:
                sys.argv = [
                    "ingest_spreadsheet.py", str(xlsx), "--html", str(html),
                    "--schools-json", str(schools_json), "--labels-json", str(labels_json),
                    "--geocode-cache", str(geocodes_json), "--no-geocode-missing",
                ]
                self.assertEqual(ingest.main(), 0)
            finally:
                sys.argv = old_argv

            schools = json.loads(schools_json.read_text(encoding="utf-8"))
            labels = json.loads(labels_json.read_text(encoding="utf-8"))
            self.assertEqual(len(schools), 1)
            self.assertEqual(schools[0]["name"], "Test School")
            self.assertEqual(schools[0]["lat"], 47.61)
            self.assertEqual(labels["cities"], [{"name": "Seattle", "lat": 47.61, "lng": -122.33, "n": 1}])

    def test_geocode_missing_uses_injected_nominatim_function(self):
        school = ingest.base_school()
        school.update({"name": "Geo School", "address": "1 Main", "city": "Seattle", "state": "WA", "country": "USA"})
        cache = {}
        original = ingest.geocode_one_nominatim
        try:
            ingest.geocode_one_nominatim = lambda s, ua: {
                "lat": 47.123456,
                "lng": -122.123456,
                "approx": False,
                "source": "fake",
                "query": ingest.geocode_query(s),
            }
            missing = ingest.geocode_missing([school], cache, "test-agent", delay=0)
        finally:
            ingest.geocode_one_nominatim = original

        self.assertEqual(missing, [])
        self.assertEqual(school["lat"], 47.123456)
        self.assertIn(ingest.geo_key(school), cache)


if __name__ == "__main__":
    unittest.main()
